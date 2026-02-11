import io
import logging
from typing import Optional

import openpyxl
import pandas as pd
import requests

from .errors import SeminarDownloaderHttpError

logger = logging.getLogger(__name__)


def get_first_row(workbook):
    sheet = workbook.active

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None and isinstance(cell_value, str) and cell_value.strip() == "Nr":
                return row

    # We couldn't find the header row, so we return 1 to get the full sheet
    return 1


class SeminarDownloader:
    session = requests.session()

    username: str
    password: str
    gliederung_id: str

    def __init__(self,
                 gliederung_id: str,
                 seminar_id: int,
                 username: str,
                 password: str,
                 user_agent: str,
                 include_non_participant_roles: bool = False,
                 include_non_standard_roles: list | None = None,
                 ) -> None:
        if include_non_standard_roles is None:
            self.include_non_standard_roles = []
        else:
            self.include_non_standard_roles = include_non_standard_roles
        self.include_non_participant_roles = include_non_participant_roles
        self._bytes = None
        self.username = username
        self.password = password
        self.gliederung_id = gliederung_id
        self.seminar_id = seminar_id
        self.headers = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "de,en-US;q=0.7,en;q=0.3",
            "Accept-Encoding": "gzip, deflate, br",
            "Referer": "https://dlrg.net/",
            "Content-Type": "application/x-www-form-urlencoded",
            "Origin": "https://dlrg.net",
            "DNT": "1",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "same-origin",
            "Sec-Fetch-User": "?1",
            "TE": "trailers",
            "User-Agent": user_agent,
        }
        self._login()

    def _login(self) -> None:
        self.session.get('https://dlrg.net')

        self.session.post(
            "https://dlrg.net",
            headers=self.headers,
            data={
                'auth[user]': self.username,
                'auth[pass]': self.password,
                'url_params': '',
            },
        )

    def _get_file(self) -> io.BytesIO:
        excel_headers = {
            **self.headers,
            'Referer': f'https://dlrg.net/apps/seminar?page=planung&action=edit&id={self.seminar_id}',
        }

        result = self.session.post(
            f'https://dlrg.net/apps/seminar?page=loadDokumente&format=pdf&edvnummer={self.gliederung_id}&id={self.seminar_id}&noheader=1',
            headers=excel_headers,
            data={
                "dokumentListeTyp": "xls",
                "dokumentListeRolleList[]": [
                    "1",
                ] + ([
                     "2",  # Leiter
                     "3",  # Referent
                     "4",  # Hospitant
                     "5",  # Gast
                     "6",  # OrgaTeam
                 ] if self.include_non_participant_roles else []
                ) + self.include_non_standard_roles,
                "dokumentListeStatusList[]": ["0"],
                "dokumentListeSortierung": "anmeldenummer",
                "dokumentListeTnstatusBestaetigtDurchTeilnehmer": "",
                "dokumentListeTnstatusBestaetigtDurchVerwalter": "",
                "dokumentListeTnstatusBestaetigtDurchGliederung": "",
                "dokumentListeTnstatusTeilgenommen": "",
                "dokumentListeTnstatusBestanden": "",
                "dokumentListeShowAllgemeineLehrgangInfos": "0",
            },
        )

        logger.info(
            f'Got status code {result.status_code} for Excel download {self.seminar_id}'
        )

        if 200 > result.status_code >= 400:
            logger.warning(f'Excel download failed with status code {result.status_code}')
            raise SeminarDownloaderHttpError(f'Excel download failed with status code {result.status_code}')

        return io.BytesIO(result.content)

    def get_data(self, *, write_file: Optional[str] = None) -> list[pd.DataFrame]:
        self._bytes = self._get_file()

        if write_file is not None:
            with open(
                    f"{write_file}.xlsx", "wb"
            ) as file:
                file.write(self._bytes.getbuffer())

        workbook = openpyxl.load_workbook(filename=self._bytes)
        row = get_first_row(workbook)

        sheets = len(workbook.sheetnames)

        return [
            pd.read_excel(self._bytes, sheet_name=0, engine='openpyxl', converters={'Plz': str}, skiprows=row - 1)
        ] + [
            pd.read_excel(self._bytes, sheet_name=i, engine='openpyxl', converters={'Plz': str})
            for i in range(1, sheets)
        ]
